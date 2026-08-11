import json
import unittest
from pathlib import Path

from openpyxl import load_workbook
import pdfplumber
import pypdfium2 as pdfium


ROOT = Path(__file__).resolve().parents[1]
DEMO = ROOT / "test_data" / "enterprise_demo"


class EnterpriseDemoTests(unittest.TestCase):
    def test_required_deliverables_exist(self):
        expected = [
            "00_采购文件_XX市信息化平台升级建设项目.pdf",
            "00_采购文件_XX市信息化平台升级建设项目.docx",
            "A_华诚科技有限公司_投标响应文件.pdf",
            "B_博远信息技术有限公司_投标响应文件.pdf",
            "C_新联科技有限公司_投标响应文件.pdf",
            "D_天远科技有限公司_投标响应文件.pdf",
            "05_电子交易与评审元数据.xlsx",
            "expected_findings.json",
        ]
        for name in expected:
            self.assertTrue((DEMO / name).is_file(), name)

    def test_bidder_pdfs_have_fixed_visual_test_pages(self):
        for code, name in (("A", "华诚科技有限公司"), ("B", "博远信息技术有限公司"), ("C", "新联科技有限公司"), ("D", "天远科技有限公司")):
            reader = pdfium.PdfDocument(str(DEMO / f"{code}_{name}_投标响应文件.pdf"))
            try:
                self.assertEqual(len(reader), 15)
            finally:
                reader.close()
        with pdfplumber.open(DEMO / "D_天远科技有限公司_投标响应文件.pdf") as d_reader:
            self.assertIn("投标人（盖章）", d_reader.pages[7].extract_text())
            self.assertIn("法定代表人（签字）", d_reader.pages[14].extract_text())

    def test_metadata_contains_combined_signals(self):
        workbook = load_workbook(DEMO / "05_电子交易与评审元数据.xlsx", data_only=False)
        rows = list(workbook["文件与网络元数据"].iter_rows(min_row=2, values_only=True))
        b, c = rows[1], rows[2]
        self.assertEqual(b[5:9], c[5:9])
        bids = [row[2] for row in workbook["投标记录"].iter_rows(min_row=2, values_only=True)]
        self.assertEqual(bids[1:4], [9_600_000, 9_800_000, 10_000_000])
        self.assertEqual(bids[2] - bids[1], bids[3] - bids[2])

    def test_expected_findings_require_human_review(self):
        payload = json.loads((DEMO / "expected_findings.json").read_text(encoding="utf-8"))
        self.assertTrue(all(item["requires_human_review"] for item in payload["expected_findings"]))
        self.assertIn("不直接认定串通投标", payload["principle"])


if __name__ == "__main__":
    unittest.main()
