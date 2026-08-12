from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


HEADERS = {
    "compliance": ["问题编号", "风险等级", "问题类型", "问题描述", "原文证据", "依据条款", "原文出处", "整改建议", "最终状态"],
    "data": ["问题编号", "风险等级", "问题类型", "问题描述", "复算或比对证据", "核验依据", "原文出处", "处理建议", "排序/一致性结论"],
    "anomaly": ["线索编号", "风险等级", "异常类型", "线索描述", "证据链", "证据来源", "涉及主体", "分析依据", "处置建议"],
}


def create_deliverable_xlsx(payload: dict, kind: str, path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "成果说明"
    _write_cover(ws, payload, kind)
    if kind == "parse":
        _write_parse_package(wb, payload["document_parse_package"])
    else:
        key = {"compliance": "compliance_issue_list", "data": "data_verification_result", "anomaly": "anomaly_warning_report"}[kind]
        rows = payload[key].get("issues", payload[key].get("warnings", []))
        ws = wb.create_sheet({"compliance": "合规问题清单", "data": "数据核验结果", "anomaly": "异常预警线索"}[kind])
        _write_issue_sheet(ws, kind, rows)
    _style_workbook(wb)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _write_cover(ws, payload: dict, kind: str) -> None:
    titles = {"parse": "结构化解析数据包", "compliance": "合规问题清单", "data": "数据核验结果", "anomaly": "异常评分预警及围串标线索报告"}
    descriptions = {
        "parse": "采购文件、响应文件、开评标资料的结构化解析成果，包含打分、汇总、废标、评审意见和候选人排序。",
        "compliance": "合规审查发现的问题点、原文证据、出处、依据及整改建议。",
        "data": "报价、评分、权重、汇总、基础字段及候选人排序的复算和一致性核验结果。",
        "anomaly": "异常评分、文件雷同、主体关联、设备网络及报价规律等风险线索；不直接认定围串标。",
    }
    section = {"compliance": "compliance_issue_list", "data": "data_verification_result", "anomaly": "anomaly_warning_report"}.get(kind)
    count = len(payload[section].get("issues", payload[section].get("warnings", []))) if section else len(payload["document_parse_package"].get("documents", []))
    ws.append([titles[kind]])
    ws.append(["项目编号", payload.get("project_id", "")])
    ws.append(["项目名称", payload.get("project_name", "")])
    ws.append(["任务编号", payload.get("task_id", "")])
    ws.append(["任务状态", payload.get("status", "")])
    ws.append(["成果记录数", count])
    ws.append(["成果说明", descriptions[kind]])
    ws.append(["使用说明", "本成果基于自动核验及人工复核状态生成；涉及待人工复核的内容不得直接作为行政或法律认定依据。"])
    ws.merge_cells("A1:B1")
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="173B76")
    ws.row_dimensions[1].height = 28


def _write_parse_package(wb: Workbook, package: dict) -> None:
    ws = wb.create_sheet("文档解析总览")
    ws.append(["文件名称", "文档类型", "页数", "解析状态", "OCR", "章节数", "表格数", "项目名称", "采购人/招标人", "代理机构"])
    for doc in package.get("documents", []):
        ws.append([doc.get("filename", ""), doc.get("document_subtype", doc.get("file_type", "")), doc.get("page_count", 0), doc.get("parse_status", ""), "是" if doc.get("ocr_applied") else "否", len(doc.get("sections", [])), len(doc.get("tables", [])), doc.get("project_name", ""), doc.get("tenderer", ""), doc.get("procurement_agency", "")])
    for title, key in [("打分明细", "score_details"), ("汇总数据", "score_summaries"), ("废标说明", "rejection_records"), ("评审意见", "evaluation_opinions"), ("候选人排序", "candidate_rankings")]:
        sheet = wb.create_sheet(title)
        items = package.get(key, [])
        if not items:
            sheet.append(["核验状态", "未形成可交付明细"])
            sheet.append(["说明", f"本批材料自动解析后未识别到可验证的{title}。这不等于原始文件一定不存在该内容；如该项属于必备资料，请回看原文或提交人工复核。"])
            sheet.append(["需人工复核", "是"])
            continue
        columns = list(dict.fromkeys(k for item in items for k in item.keys()))
        sheet.append(columns)
        for item in items:
            sheet.append([_cell(item.get(col, "")) for col in columns])


def _write_issue_sheet(ws, kind: str, rows: list[dict]) -> None:
    ws.append(HEADERS[kind])
    if not rows:
        ws.append(["—", "—", "本次未形成异常线索" if kind == "anomaly" else "本次未形成问题", "自动核验未发现具有充分证据的可输出事项。", "无", "无", "无", "保持关注并结合新增材料复核。", "passed"])
        return
    for item in rows:
        common = [item.get("issue_id", ""), item.get("risk_level", ""), item.get("issue_type", ""), item.get("description", "")]
        if kind == "compliance":
            values = common + [item.get("evidence", ""), item.get("basis", ""), item.get("source_location", ""), item.get("suggestion", ""), item.get("final_status", item.get("assessment", ""))]
        elif kind == "data":
            values = common + [item.get("evidence", ""), item.get("basis", ""), item.get("source_location", ""), item.get("suggestion", ""), item.get("assessment", item.get("final_status", ""))]
        else:
            values = common + [item.get("evidence", ""), _cell(item.get("evidence_refs", item.get("evidence_sources", ""))), _cell(item.get("related_entities", "")), item.get("basis", ""), item.get("suggestion", "")]
        ws.append([_cell(value) for value in values])


def _cell(value) -> str | int | float:
    if isinstance(value, list):
        return "\n".join(str(x) for x in value)
    if isinstance(value, dict):
        return "; ".join(f"{k}: {v}" for k, v in value.items())
    return value if isinstance(value, (str, int, float)) else str(value or "")


def _style_workbook(wb: Workbook) -> None:
    border = Border(bottom=Side(style="thin", color="D9E2F1"))
    for ws in wb.worksheets:
        if ws.title != "成果说明":
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2563EB")
        for index, column in enumerate(ws.columns, start=1):
            letter = get_column_letter(index)
            ws.column_dimensions[letter].width = min(48, max(12, max(len(str(c.value or "")) for c in column) + 2))
            for cell in column:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
